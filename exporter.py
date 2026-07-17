"""
ThreadScout - Exporter Module
===============================
Export collected data to CSV and Excel formats.
"""

from pathlib import Path

import pandas as pd
from loguru import logger
from rich.progress import Progress, SpinnerColumn, TextColumn, BarColumn

from config import COLORS, CSV_OUTPUT, EXCEL_OUTPUT
from ui import console, show_success, show_error, show_info


def _prepare_dataframe(results: list[dict]) -> pd.DataFrame:
    """
    Convert results list to a structured Pandas DataFrame.

    Args:
        results: List of result dictionaries.

    Returns:
        DataFrame with standardized columns.
    """
    columns = [
        "Threads",
        "Instagram",
        "Instagram Link",
        "Keyword",
        "Post URL",
        "Post Content",
        "Date Scraped",
    ]

    if not results:
        return pd.DataFrame(columns=columns)

    df = pd.DataFrame(results)

    # Ensure all required columns exist
    for col in columns:
        if col not in df.columns:
            df[col] = ""

    # Reorder and select only required columns
    df = df[columns]

    return df


def export_csv(results: list[dict], filepath: Path | None = None) -> bool:
    """
    Export results to a CSV file.

    Args:
        results: List of result dictionaries.
        filepath: Custom output path (defaults to config CSV_OUTPUT).

    Returns:
        True if export was successful, False otherwise.
    """
    output_path = filepath or CSV_OUTPUT

    if not results:
        show_error("Tidak ada data untuk diekspor.")
        return False

    try:
        # Ensure output directory exists
        output_path.parent.mkdir(parents=True, exist_ok=True)

        with Progress(
            SpinnerColumn(style=COLORS["pink"]),
            TextColumn(f"[{COLORS['white']}]Mengekspor CSV..."),
            BarColumn(complete_style=COLORS["purple"]),
            console=console,
            transient=True,
        ) as progress:
            task = progress.add_task("export", total=3)

            # Step 1: Prepare DataFrame
            df = _prepare_dataframe(results)
            progress.advance(task)

            # Step 2: Write CSV
            df.to_csv(output_path, index=False, encoding="utf-8-sig")
            progress.advance(task)

            # Step 3: Verify
            progress.advance(task)

        show_success(f"CSV berhasil diekspor → {output_path}")
        show_info(f"Total {len(results)} baris data")
        logger.info(f"CSV exported: {output_path} ({len(results)} rows)")
        return True

    except PermissionError:
        show_error(f"File sedang digunakan: {output_path}")
        show_info("Tutup file terlebih dahulu dan coba lagi.")
        logger.error(f"Permission denied exporting CSV: {output_path}")
        return False
    except OSError as e:
        show_error(f"Gagal mengekspor CSV: {e}")
        logger.error(f"CSV export error: {e}")
        return False


def export_excel(results: list[dict], filepath: Path | None = None) -> bool:
    """
    Export results to an Excel file with styled formatting.

    Args:
        results: List of result dictionaries.
        filepath: Custom output path (defaults to config EXCEL_OUTPUT).

    Returns:
        True if export was successful, False otherwise.
    """
    output_path = filepath or EXCEL_OUTPUT

    if not results:
        show_error("Tidak ada data untuk diekspor.")
        return False

    try:
        # Ensure output directory exists
        output_path.parent.mkdir(parents=True, exist_ok=True)

        with Progress(
            SpinnerColumn(style=COLORS["pink"]),
            TextColumn(f"[{COLORS['white']}]Mengekspor Excel..."),
            BarColumn(complete_style=COLORS["purple"]),
            console=console,
            transient=True,
        ) as progress:
            task = progress.add_task("export", total=4)

            # Step 1: Prepare DataFrame
            df = _prepare_dataframe(results)
            progress.advance(task)

            # Step 2: Create Excel writer with openpyxl engine
            with pd.ExcelWriter(
                output_path, engine="openpyxl", mode="w"
            ) as writer:
                df.to_excel(writer, index=False, sheet_name="ThreadScout Results")
                progress.advance(task)

                # Step 3: Style the workbook
                workbook = writer.book
                worksheet = writer.sheets["ThreadScout Results"]

                # Header styling
                from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

                header_font = Font(
                    name="Segoe UI",
                    bold=True,
                    color="FFFFFF",
                    size=11,
                )
                header_fill = PatternFill(
                    start_color="833AB4",
                    end_color="833AB4",
                    fill_type="solid",
                )
                header_alignment = Alignment(
                    horizontal="center",
                    vertical="center",
                    wrap_text=True,
                )
                thin_border = Border(
                    left=Side(style="thin", color="D0D0D0"),
                    right=Side(style="thin", color="D0D0D0"),
                    top=Side(style="thin", color="D0D0D0"),
                    bottom=Side(style="thin", color="D0D0D0"),
                )

                for col_num, col_name in enumerate(df.columns, 1):
                    cell = worksheet.cell(row=1, column=col_num)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                    cell.border = thin_border

                # Data cell styling
                data_font = Font(name="Segoe UI", size=10)
                data_alignment = Alignment(vertical="center", wrap_text=True)

                for row in range(2, len(df) + 2):
                    for col in range(1, len(df.columns) + 1):
                        cell = worksheet.cell(row=row, column=col)
                        cell.font = data_font
                        cell.alignment = data_alignment
                        cell.border = thin_border

                        # Alternate row coloring
                        if row % 2 == 0:
                            cell.fill = PatternFill(
                                start_color="F8F0FF",
                                end_color="F8F0FF",
                                fill_type="solid",
                            )

                # Column widths
                column_widths = {
                    "A": 18,  # Threads
                    "B": 22,  # Instagram
                    "C": 38,  # Instagram Link
                    "D": 16,  # Keyword
                    "E": 40,  # Post URL
                    "F": 50,  # Post Content
                    "G": 18,  # Date Scraped
                }

                for col_letter, width in column_widths.items():
                    worksheet.column_dimensions[col_letter].width = width

                # Freeze top row
                worksheet.freeze_panes = "A2"

                progress.advance(task)

            # Step 4: Complete
            progress.advance(task)

        show_success(f"Excel berhasil diekspor → {output_path}")
        show_info(f"Total {len(results)} baris data dengan format styling")
        logger.info(f"Excel exported: {output_path} ({len(results)} rows)")
        return True

    except PermissionError:
        show_error(f"File sedang digunakan: {output_path}")
        show_info("Tutup file terlebih dahulu dan coba lagi.")
        logger.error(f"Permission denied exporting Excel: {output_path}")
        return False
    except OSError as e:
        show_error(f"Gagal mengekspor Excel: {e}")
        logger.error(f"Excel export error: {e}")
        return False
